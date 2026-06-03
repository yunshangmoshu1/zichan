# -*- coding: utf-8 -*-
"""
将钉钉导出的 xlsx 文件转换为 JSON 格式，用于导入 Supabase。
用法: python convert_xlsx_to_json.py
输出: robots_data.json
"""

import openpyxl
import json
import sys

INPUT_FILE = "2026机器人出入库明细表.xlsx"
OUTPUT_FILE = "robots_data.json"
SHEET_NAME = "入库管理"

# 列索引 (0-based)
COL_TYPE = 0       # 机器人类型
COL_SERIAL = 1     # 机器人出厂编号
COL_STATUS = 2     # 机器人状态
COL_PERSON = 3     # 关联责任人
COL_IP = 4         # IP
COL_LOCATION = 5   # 位置更新时间（每周五更新）
COL_NOTES = 6      # 备注
COL_BORROWED = 7   # 是否借出
COL_IMAGE = 8      # 相关图片
COL_UPDATER = 11   # 更新人
COL_CREATED = 12   # 创建时间


def convert():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    # 查找工作表（处理可能的编码问题）
    ws = None
    for name in wb.sheetnames:
        if "入库" in name:
            ws = wb[name]
            print(f"找到工作表: {name}")
            break

    if ws is None:
        print(f"错误: 未找到包含'入库'的工作表。可用的工作表: {wb.sheetnames}")
        sys.exit(1)

    robots = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[COL_SERIAL]:
            continue

        robot_type = str(row[COL_TYPE] or "").strip()
        serial = str(row[COL_SERIAL] or "").strip()

        if not robot_type or not serial:
            continue

        borrowed = False
        borrowed_val = str(row[COL_BORROWED] or "").strip().upper()
        if borrowed_val in ("OK", "TRUE", "是", "1"):
            borrowed = True

        robot = {
            "type": robot_type,
            "serial": serial,
            "status": str(row[COL_STATUS] or "测试中").strip(),
            "person": str(row[COL_PERSON] or "").strip() or None,
            "ip": str(row[COL_IP] or "").strip() or None,
            "location": str(row[COL_LOCATION] or "").strip() or None,
            "notes": str(row[COL_NOTES] or "").strip() or None,
            "borrowed": borrowed,
            "image": str(row[COL_IMAGE] or "").strip() or None,
            "updater": str(row[COL_UPDATER] or "").strip() or None,
        }

        # 处理创建时间
        created = row[COL_CREATED]
        if created:
            robot["created_at"] = str(created)
        else:
            robot["created_at"] = None

        robots.append(robot)

    wb.close()

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(robots, f, ensure_ascii=False, indent=2)

    print(f"转换完成: {len(robots)} 条记录 -> {OUTPUT_FILE}")

    # 统计
    types = {}
    statuses = {}
    for r in robots:
        types[r["type"]] = types.get(r["type"], 0) + 1
        statuses[r["status"]] = statuses.get(r["status"], 0) + 1

    print("\n按类型统计:")
    for t, c in sorted(types.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}")

    print("\n按状态统计:")
    for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c}")


if __name__ == "__main__":
    convert()
